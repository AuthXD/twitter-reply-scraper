#!/usr/bin/env python3
"""Proof-log auto-tracker.

Reads Proof_Log.xlsx, and for every token you've logged (any row with a
Contract Address), pulls live market data from Dexscreener and fills in the
"what it did after" side of the receipt automatically:

  * Peak MC ($)   -> running maximum of every reading (never goes down)
  * Peak Price ($)
  * Time to Peak  -> stamped when a new peak is reached
  * What It Did After -> a regenerated "[auto ...]" summary line
  * Status        -> Pending / Hit / Big hit / Flat / Rug / dumped

It only *adds* the after-the-fact data. Your at-flag columns (timestamp, token,
entry signal, Entry MC, Entry Price) and anything you typed by hand are left
alone. The Multiple (x) column stays a live formula, so Excel recomputes it the
moment you open the file.

Stdlib only (openpyxl + urllib) — no 'requests', no API key, nothing to install
beyond openpyxl, which the pipeline already uses.

Usage:
    python proof_tracker.py                      # update the default file once
    python proof_tracker.py --file path.xlsx     # point at a specific file
    python proof_tracker.py --dry-run            # show changes, write nothing
"""
import argparse
import datetime as dt
import json
import logging
import os
import sys
import urllib.request

from openpyxl import load_workbook

DEFAULT_FILE = os.environ.get(
    "PROOF_LOG",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "Proof_Log.xlsx"),
)
SHEET = "Proof Log"
HEADER_ROW = 3
API = "https://api.dexscreener.com/latest/dex/tokens/{addr}"
AUTO = "[auto"          # marker so we only overwrite our own narrative
UA = "proof-tracker/1.0"

# 1-based column indices (must match the workbook layout)
C = dict(call=1, ts=2, token=3, ticker=4, chain=5, contract=6, signal=7,
         entry_mc=8, entry_price=9, peak_mc=10, peak_price=11, mult=12,
         time_to_peak=13, after=14, status=15, source=16, shot=17, notes=18)

log = logging.getLogger("reply_scraper.proof")


# --------------------------------------------------------------------------- #
# Market data
# --------------------------------------------------------------------------- #
def dexscreener_lookup(address, timeout=20):
    """Return {mc, price, url} for the deepest-liquidity pair, or None."""
    req = urllib.request.Request(API.format(addr=address), headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    return _best_pair(data.get("pairs") or [])


def _best_pair(pairs):
    best = None
    best_liq = -1.0
    for p in pairs:
        liq = _num((p.get("liquidity") or {}).get("usd"))
        if liq is None:
            liq = 0.0
        if liq > best_liq:
            best_liq = liq
            best = p
    if not best:
        return None
    mc = _num(best.get("marketCap")) or _num(best.get("fdv"))
    price = _num(best.get("priceUsd"))
    return {"mc": mc, "price": price, "url": best.get("url", "")}


def _num(v):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# --------------------------------------------------------------------------- #
# Status logic
# --------------------------------------------------------------------------- #
def classify(entry_mc, peak_mc, cur_mc):
    """Map the peak multiple + current drawdown to a Status label."""
    if not entry_mc or entry_mc <= 0 or not peak_mc:
        return "Pending"
    peak_mult = peak_mc / entry_mc
    cur_mult = (cur_mc / entry_mc) if cur_mc else 0.0
    if cur_mc and entry_mc and cur_mc < entry_mc * 0.15 and peak_mult < 2:
        return "Rug / dumped"
    if peak_mult >= 10:
        return "Big hit"
    if peak_mult >= 2:
        return "Hit"
    if 0.8 <= cur_mult <= 1.25:
        return "Flat"
    return "Pending"


def _fmt_money(v):
    if v is None:
        return "?"
    if v >= 1_000_000:
        return f"${v/1_000_000:.2f}M"
    if v >= 1_000:
        return f"${v/1_000:.0f}K"
    return f"${v:,.0f}"


def _fmt_mult(entry, mc):
    if not entry or entry <= 0 or mc is None:
        return "?"
    return f"{mc/entry:.1f}x"


def _fmt_duration(start, end):
    secs = max(0, int((end - start).total_seconds()))
    d, rem = divmod(secs, 86400)
    h, rem = divmod(rem, 3600)
    m, _ = divmod(rem, 60)
    if d:
        return f"{d}d {h}h"
    if h:
        return f"{h}h {m}m"
    return f"{m}m"


def _parse_ts(val):
    if isinstance(val, dt.datetime):
        return val
    if isinstance(val, str) and val.strip():
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M"):
            try:
                return dt.datetime.strptime(val.strip(), fmt)
            except ValueError:
                continue
    return None


# --------------------------------------------------------------------------- #
# Row update
# --------------------------------------------------------------------------- #
def update_row(ws, r, now, dry_run):
    def g(key):
        return ws.cell(r, C[key]).value

    contract = g("contract")
    if not contract or not str(contract).strip():
        return False
    notes = str(g("notes") or "")
    if notes.startswith("EXAMPLE"):
        return False  # leave the sample rows alone

    address = str(contract).strip()
    try:
        live = dexscreener_lookup(address)
    except Exception as exc:
        log.warning("row %s (%s): lookup failed: %s", r, g("ticker") or address[:8], exc)
        return False
    if not live or live["mc"] is None:
        log.info("row %s (%s): no live pair/marketcap yet", r, g("ticker") or address[:8])
        return False

    entry_mc = _num(g("entry_mc"))
    prev_peak = _num(g("peak_mc")) or 0.0
    cur_mc = live["mc"]
    new_peak = max(prev_peak, cur_mc)
    peak_advanced = new_peak > prev_peak + 1e-9

    changes = {}
    changes["peak_mc"] = new_peak
    if live["price"] is not None:
        prev_pp = _num(g("peak_price")) or 0.0
        changes["peak_price"] = max(prev_pp, live["price"])

    ts = _parse_ts(g("ts"))
    if peak_advanced and ts:
        changes["time_to_peak"] = _fmt_duration(ts, now)

    stamp = now.strftime("%Y-%m-%d %H:%MZ")
    changes["after"] = (
        f"[auto {stamp}] peak {_fmt_money(new_peak)} MC "
        f"({_fmt_mult(entry_mc, new_peak)} from entry); "
        f"now {_fmt_money(cur_mc)} ({_fmt_mult(entry_mc, cur_mc)})."
    )

    new_status = classify(entry_mc, new_peak, cur_mc)
    cur_status = str(g("status") or "").strip()
    if cur_status in ("", "Pending", "Flat", "Hit", "Big hit", "Rug / dumped"):
        changes["status"] = new_status

    # Only overwrite the narrative if it's empty or one of ours.
    existing_after = str(g("after") or "")
    if existing_after and not existing_after.startswith(AUTO):
        changes.pop("after", None)

    label = g("ticker") or g("token") or address[:8]
    if dry_run:
        log.info("row %s (%s) WOULD SET: %s", r, label,
                 {k: changes[k] for k in changes})
        return False

    for key, val in changes.items():
        ws.cell(r, C[key]).value = val
    log.info("row %s (%s): peak %s status %s", r, label,
             _fmt_money(new_peak), changes.get("status", cur_status))
    return True


def run(path, dry_run=False):
    if not os.path.exists(path):
        log.error("proof log not found: %s", path)
        return 1
    wb = load_workbook(path)  # keep formulas (Multiple column stays live)
    if SHEET not in wb.sheetnames:
        log.error("sheet %r not in %s", SHEET, path)
        return 1
    ws = wb[SHEET]
    # utcnow() is deprecated from 3.12; the sheet stamps are naive UTC, so drop
    # the tzinfo after converting rather than reading local time by accident.
    now = dt.datetime.now(dt.timezone.utc).replace(tzinfo=None)

    touched = 0
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        try:
            if update_row(ws, r, now, dry_run):
                touched += 1
        except Exception as exc:  # never let one row kill the run
            log.warning("row %s errored, skipping: %s", r, exc)

    if touched and not dry_run:
        wb.save(path)
    log.info("done: %d row(s) updated%s", touched, " (dry-run, not saved)" if dry_run else "")
    return 0


def main(argv=None):
    p = argparse.ArgumentParser(description="Auto-update the proof log from live market data.")
    p.add_argument("--file", default=DEFAULT_FILE, help="path to Proof_Log.xlsx")
    p.add_argument("--dry-run", action="store_true", help="print changes without writing")
    args = p.parse_args(argv)
    logging.basicConfig(
        level=logging.INFO, stream=sys.stdout,
        format="%(asctime)s %(levelname)s %(message)s")
    return run(args.file, dry_run=args.dry_run)


if __name__ == "__main__":
    raise SystemExit(main())
