#!/usr/bin/env python3
"""Build Leads.xlsx from qualified.csv (ranked, formatted, clickable).

Observed data only. The sheet reports what was posted, by whom and where — it
does not draft copy, suggest an angle, or track who has been contacted. Those
belong to whoever runs outreach, not to a scraper's output: a suggestion column
invites the sheet to be worked top-down as a queue, and a status column makes
that the intended use. Collection and outreach are kept separate on purpose.
"""
import csv, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = sys.argv[1] if len(sys.argv) > 1 else "qualified.csv"
OUT = sys.argv[2] if len(sys.argv) > 2 else "Leads.xlsx"
FONT, NAVY, ACCENT, HI, MED = "Arial", "0B1F3A", "1F6FEB", "FDE7E9", "FFF6D6"

def main():
    if not os.path.exists(SRC):
        print(f"{SRC} not found", file=sys.stderr); return 1
    rows = list(csv.DictReader(open(SRC, encoding="utf-8")))
    rows.sort(key=lambda r: (-int(r["score"]), r["intent"]))
    wb = Workbook(); ws = wb.active; ws.title = "Leads"
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)
    center = Alignment(vertical="top", horizontal="center", wrap_text=True)
    cols = [("#",5),("Intent",9),("Score",7),("Handle",16),("Theme",22),("What they said",52),
            ("Matched",14),("Under caller",15),("Date",11)]
    ws.merge_cells("A1:I1"); t = ws["A1"]
    t.value = "QUALIFIED LEADS   |   people posting the pain, ranked by intent"
    t.font = Font(name=FONT, bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=ACCENT)
    t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:I2"); s = ws["A2"]
    s.value = ("Observed data only — what was posted, by whom, and where. Handle links to their X "
               "profile; 'What they said' links to the reply. Rebuilt each time the scraper runs.")
    s.font = Font(name=FONT, italic=True, size=9, color="44506A")
    s.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
    ws.row_dimensions[2].height = 26
    HR = 3
    for i, (name, w) in enumerate(cols, 1):
        c = ws.cell(HR, i, name); c.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=NAVY); c.border = border; c.alignment = center
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[HR].height = 26
    r = HR + 1
    for i, row in enumerate(rows, 1):
        handle = row["handle"].lstrip("@")
        fill = HI if row["intent"] == "HIGH" else MED
        vals = [i, row["intent"], int(row["score"]), "@"+handle, row["theme"], row["quote"],
                row["matched_phrase"], row["caller"], row["created_at"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci, v); c.border = border; c.font = Font(name=FONT, size=10)
            c.alignment = center if ci in (1,2,3,9) else wrap
            if ci in (1,2,3): c.fill = PatternFill("solid", fgColor=fill)
        ws.cell(r,2).font = Font(name=FONT, size=10, bold=True)
        hc = ws.cell(r,4); hc.hyperlink = row["profile_url"]
        hc.font = Font(name=FONT, size=10, color="1F6FEB", underline="single")
        if row.get("source_post_url"):
            ws.cell(r,6).hyperlink = row["source_post_url"]
        r += 1
    last = max(r-1, HR+1)
    ws.freeze_panes = "A4"; ws.auto_filter.ref = f"A{HR}:I{last}"
    wb.save(OUT)
    print(f"wrote {OUT} with {len(rows)} lead(s)")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
