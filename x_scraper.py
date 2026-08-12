"""
X (Twitter) reply/search scraper — v2
=====================================
Mines reply sections / search results and adds judgment:
  * scores each hit HIGH / MED / LOW (pain weight + engagement + is-it-a-question)
  * filters to genuine retail (skips mega-influencers and bot/throwaway accounts)
  * auto-drafts a humanized reply per pain type (personalize before posting)
  * dedupes against handles already in your X Targets sheet + a seen-log
  * writes a CSV and can append straight into the lead workbook

Built on twscrape (MIT): https://github.com/vladkens/twscrape
Edit behaviour in scraper_config.py — not here.

SETUP (local, once):
  1. pip install twscrape openpyxl
  2. Add a BURNER X account (not your main; multi-account = X ToS risk + ban risk):
       twscrape add_accounts accounts.txt username:password:email:email_pw:_:cookies
  3. Run:
       python x_scraper.py --mode both --limit 40
       python x_scraper.py --mode search --to-xlsx Week1_Targets.xlsx
"""

import argparse
import asyncio
import csv
import json
import os
import re
from datetime import datetime, timezone, timedelta

import scraper_config as cfg

try:
    from twscrape import API
except ImportError:  # pragma: no cover
    API = None

SEEN_FILE = "seen_handles.json"
OUTPUT_CSV = "x_targets_scraped.csv"
COMPILED = [(re.compile(p, re.I), theme, w) for p, theme, w in cfg.LANGUAGE_PATTERNS]


def norm_handle(handle):
    """Canonical handle key: no leading @, lowercased.

    Config lists handles bare ("somecaller") while scraped rows carry them with
    an @ ("@somecaller"). Comparing the two forms directly never matched, so the
    callers themselves were never actually excluded from their own lead list.
    """
    return str(handle or "").strip().lstrip("@").lower()


EXCLUDE = {norm_handle(h) for h in cfg.CALLERS} | {norm_handle(h) for h in cfg.EXCLUDE_EXTRA}


# ---------------------------------------------------------------- classify/score
def classify(text):
    """Return (theme, matched_snippet, weight) for the strongest match, or None."""
    best = None
    for rx, theme, w in COMPILED:
        m = rx.search(text or "")
        if m and (best is None or w > best[2]):
            best = (theme, m.group(0), w)
    return best


def score(weight, text, likes):
    s = weight
    if "?" in (text or ""):
        s += 2  # asking for help = higher intent
    if (likes or 0) >= 5:
        s += 1
    if (likes or 0) >= 50:
        s += 1
    return s


def bucket(s):
    if s >= cfg.SCORE_BUCKETS["HIGH"]:
        return "HIGH"
    if s >= cfg.SCORE_BUCKETS["MED"]:
        return "MED"
    return "LOW"


# ---------------------------------------------------------------- retail filter
def passes_filters(tw):
    u = getattr(tw, "user", None)
    if not u:
        return False
    f = cfg.FILTERS
    followers = getattr(u, "followersCount", 0) or 0
    if followers < f["follower_min"] or followers > f["follower_max"]:
        return False
    if f["skip_verified_org"] and getattr(u, "verifiedType", "") == "Business":
        return False
    if f["require_recent_days"]:
        d = getattr(tw, "date", None)
        if d and d < datetime.now(timezone.utc) - timedelta(days=f["require_recent_days"]):
            return False
    return True


# ---------------------------------------------------------------- row builder
def build_row(tw, theme, snippet, weight, mode):
    u = getattr(tw, "user", None)
    handle = getattr(u, "username", "") if u else ""
    text = (getattr(tw, "rawContent", "") or "").replace("\n", " ").strip()
    likes = getattr(tw, "likeCount", 0)
    s = score(weight, text, likes)
    return {
        "handle": "@" + handle,
        "profile_url": f"https://x.com/{handle}",
        "source_post_url": getattr(tw, "url", ""),
        "date": getattr(tw, "date", ""),
        "followers": getattr(u, "followersCount", None) if u else None,
        "intent_quote": text,
        "matched_phrase": snippet,
        "pain_category": theme,
        "intent": bucket(s),
        "score": s,
        "draft_reply": cfg.REPLY_TEMPLATES.get(theme, ""),
        "found_via": mode,
    }


# ---------------------------------------------------------------- scrape modes
async def from_search(api, limit):
    out = []
    for q in cfg.SEARCH_QUERIES:
        async for tw in api.search(q, limit=limit, kv={"product": "Latest"}):
            hit = classify(getattr(tw, "rawContent", ""))
            if hit and passes_filters(tw):
                out.append(build_row(tw, hit[0], hit[1], hit[2], f"search:{q[:28]}"))
    return out


async def from_replies(api, tweets_per_caller, replies_per_tweet):
    out = []
    for handle in cfg.CALLERS:
        try:
            user = await api.user_by_login(handle)
        except Exception as e:  # noqa
            print(f"  ! skip @{handle}: {e}")
            continue
        if not user:
            continue
        async for tw in api.user_tweets(user.id, limit=tweets_per_caller):
            async for rep in api.tweet_replies(tw.id, limit=replies_per_tweet):
                hit = classify(getattr(rep, "rawContent", ""))
                if hit and passes_filters(rep):
                    out.append(build_row(rep, hit[0], hit[1], hit[2], f"reply:@{handle}"))
    return out


# ---------------------------------------------------------------- dedupe + persist
def load_seen():
    # Normalised on read, so seen-logs written by older versions still match.
    if os.path.exists(SEEN_FILE):
        with open(SEEN_FILE) as f:
            return {norm_handle(h) for h in json.load(f)}
    return set()


def save_seen(seen):
    with open(SEEN_FILE, "w") as f:
        json.dump(sorted(seen), f, indent=0)


def existing_xlsx_handles(path):
    """Handles already logged in the X Targets tab, so we never resurface them."""
    if not path or not os.path.exists(path):
        return set()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        if "X Targets" not in wb.sheetnames:
            return set()
        ws = wb["X Targets"]
        out = set()
        for row in ws.iter_rows(min_col=2, max_col=2):
            v = row[0].value
            if v and str(v).strip():
                out.add(norm_handle(v))
        return out
    except Exception:
        return set()


def dedupe(rows, already):
    seen, keep = set(already), []
    for r in sorted(rows, key=lambda x: x["score"], reverse=True):
        h = norm_handle(r["handle"])
        if h in EXCLUDE or h in seen:
            continue
        seen.add(h)
        keep.append(r)
    return keep


# ---------------------------------------------------------------- xlsx append
def append_to_xlsx(rows, path):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = load_workbook(path)
    ws = wb["X Targets"]
    fills = {"HIGH": "D8F0D8", "MED": "FFF3CC"}
    r = ws.max_row + 1
    start_num = r - 4  # header is row 4 in the Week-1 sheet
    for i, row in enumerate(rows):
        vals = [start_num + i, row["handle"], row["profile_url"], row["source_post_url"],
                str(row["date"])[:10], row["intent_quote"], row["pain_category"],
                row["intent"], row["draft_reply"], ""]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if ci == 8 and row["intent"] in fills:
                c.fill = PatternFill("solid", start_color=fills[row["intent"]])
                c.alignment = Alignment(horizontal="center", vertical="top")
        r += 1
    wb.save(path)


# ---------------------------------------------------------------- run
async def run(mode, limit, to_xlsx):
    if API is None:
        raise SystemExit("twscrape not installed. Run: pip install twscrape")
    api = API()
    if not await api.pool.accounts_info():
        raise SystemExit("No accounts in pool. Add a burner account first (see setup at top of file).")

    rows = []
    if mode in ("search", "both"):
        print("Running search queries...")
        rows += await from_search(api, limit)
    if mode in ("replies", "both"):
        print("Mining caller replies...")
        rows += await from_replies(api, tweets_per_caller=5, replies_per_tweet=limit)

    already = load_seen() | existing_xlsx_handles(to_xlsx)
    rows = dedupe(rows, already)

    fields = ["handle", "profile_url", "source_post_url", "date", "followers", "intent_quote",
              "matched_phrase", "pain_category", "intent", "score", "draft_reply", "found_via"]
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)

    if to_xlsx:
        append_to_xlsx(rows, to_xlsx)
        print(f"Appended {len(rows)} rows to {to_xlsx} (X Targets tab).")

    save_seen(load_seen() | {norm_handle(r["handle"]) for r in rows})
    hi = sum(1 for r in rows if r["intent"] == "HIGH")
    print(f"Done. {len(rows)} new targets ({hi} HIGH) -> {OUTPUT_CSV}  "
          f"({datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC})")


def main():
    ap = argparse.ArgumentParser(description="X reply/search scraper v2")
    ap.add_argument("--mode", choices=["search", "replies", "both"], default="both")
    ap.add_argument("--limit", type=int, default=40, help="per-query / per-tweet cap")
    ap.add_argument("--to-xlsx", default=None, help="append results into this workbook's X Targets tab")
    args = ap.parse_args()
    asyncio.run(run(args.mode, args.limit, args.to_xlsx))


if __name__ == "__main__":
    main()
